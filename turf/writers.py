'''
writers.py: output-writing utilities for Turf.

Authors
-------

Michael Hucka <mhucka@caltech.edu> -- Caltech Library

Copyright
---------

Copyright (c) 2018 by the California Institute of Technology.  This code is
open-source software released under a 3-clause BSD license.  Please see the
file "LICENSE" for more information.
'''

import csv
import os
import openpyxl
from   openpyxl.styles import Font
from   openpyxl.utils import get_column_letter
from   openpyxl.worksheet.write_only import WriteOnlyCell
import sys

try:
    thisdir = os.path.dirname(os.path.abspath(__file__))
    sys.path.append(os.path.join(thisdir, '../..'))
except:
    sys.path.append('../..')

import turf
from turf.messages import color, msg

# NOTE: to turn on debugging, make sure python -O was *not* used to start
# python, then set the logging level to DEBUG *before* loading this module.
# Conversely, to optimize out all the debugging code, use python -O or -OO
# and everything inside "if __debug__" blocks will be entirely compiled out.
if __debug__:
    import logging
    logging.basicConfig(level = logging.INFO)
    logger = logging.getLogger('turf')
    def log(s, *other_args): logger.debug('writers: ' + s.format(*other_args))


# Global constants.
# .............................................................................

_NUM_URLS = 5


# Main module code.
# ......................................................................

# Results are assumed to be a list of the form
#   (id, [UrlData, UrlData, ...])
# where "UrlData" is the UrlData structure retured by urlup for each
# URL found in field 856 (if any are found) for the MARC XML record.

def write_results(filename, results, include_unchanged, all):
    # Call on appropriate functions for the desired output format.
    name, extension = os.path.splitext(filename)
    if extension.lower() == '.csv':
        write_csv(filename, results, include_unchanged, all)
    else:
        write_xls(filename, results, include_unchanged, all)


def write_csv(filename, tind_results, include_unchanged, all):
    file = open(filename, 'w', newline='')

    # Write the header row.
    text = 'TIND record id'
    for i in range(1, _NUM_URLS + 1):
        text += ',Original URL {},Final URL {}'.format(i, i)
    file.write(text + '\n')
    csvwriter = csv.writer(file, delimiter=',')
    try:
        for item in tind_results:
            if not item:
                if __debug__: log('no data -- stopping')
                break
            if not item.url_data and not all:
                if __debug__: log('no URLs for {} -- not saving'.format(item.id))
                continue
            if not contains_changed_urls(item.url_data) and not (include_unchanged or all):
                if __debug__: log('URLs unchanged for {} -- skipping'.format(item.id))
                continue
            row = [item.id]
            if __debug__: log('writing row for {}'.format(item.id))
            for url_data in item.url_data:
                row.append(url_data.original)
                if url_data.error:
                    row.append('(error: {})'.format(url_data.error))
                else:
                    row.append(url_data.final or '')
            if item.url_data or all:
                csvwriter.writerow(row)
                file.flush()
    except KeyboardInterrupt:
        msg('Interrupted -- closing "{}" and exiting'.format(filename))
    except Exception:
        raise
    finally:
        file.close()


def write_xls(filename, tind_results, include_unchanged, all):
    # Create some things we reuse below.
    bold_style = Font(bold = True, underline = "single")
    hyperlink_style = Font(underline='single', color='0563C1')
    error_style = Font(color='aa2222')

    # Create a sheet in a new workbook and give it a distinctive style.
    wb = openpyxl.Workbook(write_only = True)
    sheet = wb.create_sheet()
    sheet.title = 'Results'
    sheet.sheet_properties.tabColor = 'f7ba0b'

    # Set the widths of the different columngs to something more convenient.
    column = get_column_letter(1)
    sheet.column_dimensions[column].width = 15
    for idx in range(2, _NUM_URLS*2 + 2):
        column = get_column_letter(idx)
        sheet.column_dimensions[column].width = 80

    # Set the headings and format them a little bit.
    cell1 = WriteOnlyCell(sheet, value = 'TIND Identifier')
    cell1.font = bold_style
    row = [cell1]
    for i in range(1, _NUM_URLS + 1):
        cell = WriteOnlyCell(sheet, value = 'Original URL #{}'.format(i))
        cell.font = bold_style
        row.append(cell)
        cell = WriteOnlyCell(sheet, value = 'Final URL #{}'.format(i))
        cell.font = bold_style
        row.append(cell)

    # Write the header row.
    sheet.append(row)

    # Now create the data rows.
    try:
        for row_number, item in enumerate(tind_results, 2):
            if not item:
                if __debug__: log('no data -- stopping')
                break
            if not item.url_data and not all:
                if __debug__: log('no URLs for {} -- not saving'.format(item.id))
                continue
            if not contains_changed_urls(item.url_data) and not (include_unchanged or all):
                if __debug__: log('URLs unchanged for {} -- skipping'.format(item.id))
                continue
            if __debug__: log('writing row {}'.format(row_number))
            cell = WriteOnlyCell(sheet, value = item.id)
            cell.value = hyperlink(tind_entry_url(item.id), item.id)
            cell.font = hyperlink_style
            row = [cell]
            for url_data in item.url_data:
                cell = WriteOnlyCell(sheet, value = url_data.original)
                cell.value = hyperlink(url_data.original)
                cell.font = hyperlink_style
                row.append(cell)
                if url_data.error:
                    cell = WriteOnlyCell(sheet, value = '(error: {})'.format(url_data.error))
                    cell.font = error_style
                else:
                    cell = WriteOnlyCell(sheet, value = url_data.final)
                    cell.value = hyperlink(url_data.final or '')
                    cell.font = hyperlink_style
                row.append(cell)
            sheet.append(row)
    except KeyboardInterrupt:
        msg('Interrupted -- closing "{}" and exiting'.format(filename))
    except Exception:
        raise
    finally:
        wb.save(filename = filename)


# Miscellaneous utilities.
# ......................................................................

def only_with_urls(results):
    return [r for r in results if r[1]]


def tind_entry_url(tind_id):
    return 'https://caltech.TIND.io/record/{}'.format(tind_id)


def hyperlink(url, text = None):
    return '=HYPERLINK("{}", "{}")'.format(url, text or url)


def contains_changed_urls(url_data):
    return any(item.original != item.final for item in url_data if item)


# For Emacs users
# ......................................................................
# Local Variables:
# mode: python
# python-indent-offset: 4
# End:
