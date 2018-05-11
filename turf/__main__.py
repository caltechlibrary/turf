'''
__main__: main command-line interface to turf

Authors
-------

Michael Hucka <mhucka@caltech.edu>

Copyright
---------

Copyright (c) 2018 by the California Institute of Technology.  This code is
open-source software.  Please see the file "LICENSE" for more information.
'''

import csv
import os
import plac
import sys
try:
    from termcolor import colored
except:
    pass

import turf
from turf import entries_from_file, entries_from_search
from turf.messages import msg, color


# Main program.
# ......................................................................

@plac.annotations(
    all       = ('write all entries, not only those with URLs',       'flag',   'a'),
    input     = ('read MARC from file instead of searching tind.io',  'option', 'i'),
    max       = ('retrieve at most this many results (default: all)', 'option', 'm'),
    output    = ('write output to the given file',                    'option', 'o'),
    quiet     = ('do not print any messages while working',           'flag',   'q'),
    unchanged = ("write entries with URLs even if they're unchanged", 'flag',   'u'),
    no_color  = ('do not color-code terminal output',                 'flag',   'C'),
    version   = ('print version info and exit',                       'flag',   'V'),
    search    = 'complete search URL (default: none)',
)

def main(all=False, input=None, output=None, quiet=False, max=None,
         unchanged=False, no_color=False, version=False, *search):
    '''Look for caltech.tind.io records containing URLs and return updated URLs.
If given a search query, it should be a complete search url as would be typed
into a web browser.  If given a file, it should be in MARC XML format.

By default, it writes out only entries that have URLs in MARC field 856, and
then only those whose URLs are found to dereference to a different URL after
following it.  If given the -a flag, it will write out all entries, even if
they have no URLs.  If given the -u flag, it will write out entries with URLs
even if the URLs are unchanged after dereferencing.

If given the -m option, it will only fetch and process that many results.
(The default is to process all of them.)

If given an output file, the results will be written to the file.  The format
of the file will be deduced from the file name extension (.csv or .xlsx).
In the absence of a file name extension, it will default to XLS format.
If not given an output file, the results will only be printed to the terminal.
'''

    # Our defaults are to do things like color the output, which means the
    # command line flags make more sense as negated values (e.g., "nocolor").
    # Dealing with negated variables is confusing, so turn them around here.
    colorize = 'termcolor' in sys.modules and not no_color

    # Process arguments.
    if version:
        print_version()
        sys.exit()
    if input and search:
        raise SystemExit(color('Cannot use a file and search string simultaneously',
                               'error', colorize))
    if not input and not search:
        raise SystemExit(color('Must provide either a file or a search term',
                               'error', colorize))
    if input and not input.endswith('.xml'):
        raise SystemExit(color('"{}" does not appear to be an XML file'
                               .format(input), 'error', colorize))
    if max and not quiet:
        max = int(max)
        msg('Will stop after getting {} records'.format(max), 'info', colorize)
    if not output and not quiet:
        msg("No output file specified; results won't be saved.", 'warn', colorize)
    elif not quiet:
        msg('Output will be written to {}'.format(output), 'info', colorize)
    if output:
        name, extension = os.path.splitext(output)
        if extension and extension.lower() not in ['.csv', '.xlsx']:
            raise SystemExit(color('"{}" has an unrecognized file extension'.format(output),
                                   'error', colorize))
        elif not extension:
            msg('"{}" has no name extension; defaulting to xls'.format(output),
                'warn', colorize)

    # Let's do this thing.
    results = []
    try:
        if input:
            file = None
            if os.path.exists(input):
                file = input
            elif os.path.exists(os.path.join(os.getcwd(), input)):
                file = os.path.join(os.getcwd(), input)
            else:
                raise SystemExit(color('Cannot find file "{}"'.format(input),
                                       'error', colorize))
            if not quiet:
                msg('Reading MARC XML from {}'.format(file), 'info', colorize)
            results = entries_from_file(file, max, unchanged, colorize, quiet)
        else:
            results = entries_from_search(search[0], max, unchanged, colorize, quiet)
    except Exception as e:
        msg('Exception encountered: {}'.format(e))
    finally:
        if not results:
            msg('No results returned.', 'warn', colorize)
        elif output:
            if not quiet:
                msg('Writing CSV file {}'.format(output), 'info', colorize)
            write_results(output, results, all)
        if not quiet:
            msg('Done.', 'info', colorize)


# Miscellaneous utilities.
# ......................................................................

def print_version():
    print('{} version {}'.format(turf.__title__, turf.__version__))
    print('Author: {}'.format(turf.__author__))
    print('URL: {}'.format(turf.__url__))
    print('License: {}'.format(turf.__license__))


def write_results(filename, results, all):
    if not all:
        results = only_with_urls(results)
    name, extension = os.path.splitext(filename)
    if extension.lower() == '.csv':
        write_csv(filename, results)
    else:
        write_xls(filename, results)


def write_csv(filename, results):
    with open(filename, 'w', newline='') as out:
        out.write('TIND record id,Original URL,Updated URL\n')
        csvwriter = csv.writer(out, delimiter=',')
        csvwriter.writerows(results)


def write_xls(filename, results):
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    # Create some things we reuse below.
    bold_font = Font(bold = True, underline = "single")
    hyperlink_style = Font(underline='single', color='0563C1')

    # Create a sheet in a new workbook and give it a distinctive style.
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Results'
    sheet.sheet_properties.tabColor = 'f7ba0b'

    # Set the headings and format them a little bit.
    sheet.append(['TIND record id', 'Original URL', 'Updated URL'])
    for cell in sheet["1:1"]:
        cell.font = bold_font

    # Set the widths of the different columngs to something more convenient.
    col_letter = get_column_letter(1)
    sheet.column_dimensions[col_letter].width = 15
    col_letter = get_column_letter(2)
    sheet.column_dimensions[col_letter].width = 100
    col_letter = get_column_letter(3)
    sheet.column_dimensions[col_letter].width = 100
    for row, data in enumerate(results, 2):
        tind_id = data[0]
        sheet.cell(row, 1).value = tind_id
        sheet.cell(row, 1).hyperlink = tind_entry_link(tind_id)
        sheet.cell(row, 1).font = hyperlink_style

        for col, value in enumerate(data[1:], 1):
            sheet.cell(row, col + 1).value = data[col]
            sheet.cell(row, col + 1).hyperlink = data[col]
            sheet.cell(row, col + 1).font = hyperlink_style
    wb.save(filename = filename)


def only_with_urls(results):
    return [r for r in results if r[1]]


def tind_entry_link(tind_id):
    return 'https://caltech.TIND.io/record/{}'.format(tind_id)


# Main entry point.
# ......................................................................
# The following allows users to invoke this using "python3 -m turf".

if __name__ == '__main__':
    plac.call(main)


# For Emacs users
# ......................................................................
# Local Variables:
# mode: python
# python-indent-offset: 4
# End:
